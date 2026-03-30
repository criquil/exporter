from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


SECTION_MARKERS = {"Pre-Caso", "Caso", "Post-Caso"}
NOISE_PREFIXES = ("Page ",)
DEFAULT_TOSCA_DIR = Path("tosca")
DEFAULT_OUTPUT_DIR = Path("exporter_helpers/out")


@dataclass
class RowData:
    row_number: int
    values: dict[int, str]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def row_to_values(row: tuple[Any, ...]) -> dict[int, str]:
    values: dict[int, str] = {}
    for idx, cell in enumerate(row, start=1):
        text = normalize_text(cell)
        if text:
            values[idx] = text
    return values


def is_noise(values: dict[int, str]) -> bool:
    if not values:
        return True
    joined = " | ".join(values.values())
    return any(prefix in joined for prefix in NOISE_PREFIXES)


def extract_variables(rows: list[RowData]) -> list[str]:
    tokens: set[str] = set()
    pattern = re.compile(r"\{(?:CP|B)\[[^\]]+\]\}")
    for row in rows:
        for value in row.values.values():
            tokens.update(pattern.findall(value))
    return sorted(tokens)


def parse_sections(rows: list[RowData]) -> dict[str, list[RowData]]:
    current = "SinSeccion"
    sections: dict[str, list[RowData]] = {"SinSeccion": []}

    for row in rows:
        marker = row.values.get(4, "")
        if marker in SECTION_MARKERS:
            current = marker
            sections.setdefault(current, [])
            continue
        sections.setdefault(current, []).append(row)

    return sections


def build_payload(excel_path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    worksheet = workbook.worksheets[0]

    rows: list[RowData] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = row_to_values(row)
        if is_noise(values):
            continue
        rows.append(RowData(row_number=row_number, values=values))

    sections = parse_sections(rows)
    variables = extract_variables(rows)

    serializable_sections: dict[str, list[dict[str, Any]]] = {}
    for key, section_rows in sections.items():
        serializable_sections[key] = [
            {"row": item.row_number, "values": item.values} for item in section_rows
        ]

    return {
        "source": str(excel_path),
        "sheet": worksheet.title,
        "rows_total": worksheet.max_row,
        "sections": serializable_sections,
        "variables": variables,
    }


def sanitize_name(filename: str) -> str:
    """Convierte un nombre de archivo en un slug seguro para la salida JSON.

    Ejemplo: 'Caso 999 SCACS.xlsx' → 'caso_999_scacs'
    """
    stem = Path(filename).stem
    slug = unicodedata.normalize("NFKD", stem).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^\w\s-]", "", slug).strip().lower()
    slug = re.sub(r"[\s-]+", "_", slug)
    return slug


def extract_case_id(filename: str) -> str | None:
    """Intenta extraer un ID numérico del nombre del archivo (ej. 'Caso 999 SCACS.xlsx' → '999')."""
    match = re.search(r"(\d+)", Path(filename).stem)
    return match.group(1) if match else None


def discover_excel_files(directory: Path) -> list[Path]:
    """Devuelve todos los archivos .xlsx dentro del directorio dado."""
    if not directory.is_dir():
        raise FileNotFoundError(f"No existe el directorio: {directory}")
    files = sorted(directory.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No se encontraron archivos .xlsx en: {directory}")
    return files


def resolve_output_path(excel_path: Path, output_dir: Path) -> Path:
    """Genera la ruta de salida JSON a partir del nombre del Excel."""
    slug = sanitize_name(excel_path.name)
    return output_dir / f"{slug}_excel.json"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrae estructura utilizable de un Excel exportado desde Tosca."
    )
    parser.add_argument("--input", default=None, help="Ruta del .xlsx de Tosca (opcional si se usa --dir)")
    parser.add_argument(
        "--output",
        default=None,
        help="Ruta de salida JSON (opcional, se auto-genera a partir del nombre del input)",
    )
    parser.add_argument(
        "--dir",
        default=None,
        help=f"Directorio donde buscar archivos .xlsx (default: {DEFAULT_TOSCA_DIR})",
    )
    args = parser.parse_args()

    # --- Resolver lista de archivos a procesar ---
    if args.input:
        excel_files = [Path(args.input)]
    else:
        scan_dir = Path(args.dir) if args.dir else DEFAULT_TOSCA_DIR
        excel_files = discover_excel_files(scan_dir)

    output_dir = DEFAULT_OUTPUT_DIR

    for excel_path in excel_files:
        if not excel_path.exists():
            raise FileNotFoundError(f"No existe el archivo: {excel_path}")

        # Si se dio --output explícito y es un solo archivo, usarlo
        if args.output and len(excel_files) == 1:
            out_path = Path(args.output)
        else:
            out_path = resolve_output_path(excel_path, output_dir)

        payload = build_payload(excel_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        case_id = extract_case_id(excel_path.name) or "?"
        print(f"[OK] {excel_path.name}  →  {out_path}  (caso {case_id})")


if __name__ == "__main__":
    main()
